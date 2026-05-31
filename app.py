import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime
import traceback

# -----------------------------------------------------------------------------
# CONSTANTS & CONFIGURATION
# -----------------------------------------------------------------------------
MASTER_FILE_PATH = "data/Master_Directory.csv"

# Stylings
FONT_NAME = "Arial"
COLOR_NAVY = "1F3864"
COLOR_LIGHT_BLUE = "D6E4F0"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_GREY = "F2F2F2"
COLOR_SUBTOTAL = "D9E1F2"
COLOR_ALERT_FILL = "FFCCCC"
COLOR_ALERT_FONT = "C00000"

# Borders
border_thin_gray = Side(style='thin', color='AAAAAA')
border_medium_gray = Side(style='medium', color='888888')

data_cell_border = Border(left=border_thin_gray, right=border_thin_gray, top=border_thin_gray, bottom=border_thin_gray)
header_cell_border = Border(left=border_medium_gray, right=border_medium_gray, top=border_medium_gray, bottom=border_medium_gray)

# Column Widths Maps
COL_WIDTHS_1235 = {
    'A': 6, 'B': 20, 'C': 20, 'D': 22, 'E': 9, 'F': 10, 'G': 9, 'H': 10, 'I': 9,
    'J': 10, 'K': 9, 'L': 11, 'M': 9, 'N': 14, 'O': 12
}
COL_WIDTHS_4 = {
    'A': 6, 'B': 22, 'C': 13, 'D': 10, 'E': 9, 'F': 10, 'G': 9, 'H': 12,
    'I': 13, 'J': 10, 'K': 9, 'L': 10, 'M': 9, 'N': 12
}

# -----------------------------------------------------------------------------
# HELPER FUNCTIONS FOR CALCULATIONS
# -----------------------------------------------------------------------------
def safe_div_model_b(num_del, num_red, num_ret, den_rec):
    try:
        den = float(den_rec)
        if den == 0 or pd.isna(den):
            return "-"
        val = ((float(num_del or 0) + float(num_red or 0) + float(num_ret or 0)) / den) * 100.0
        return round(val, 1)
    except:
        return "-"

def safe_div_model_a(num_dss, den_pdm):
    try:
        den = float(den_pdm)
        if den == 0 or pd.isna(den):
            return "-"
        val = (float(num_dss or 0) / den) * 100.0
        return round(val, 1)
    except:
        return "-"

def format_cell_value(val):
    if val == "-":
        return "-"
    try:
        return float(val)
    except:
        return str(val)

# -----------------------------------------------------------------------------
# WORKBOOK GENERATION ENGINE
# -----------------------------------------------------------------------------
def generate_excel_report(df_final, counts_map, daily_date, cumulative_range, dss_range):
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # -------------------------------------------------------------------------
    # STYLING WRAPPER FUNCTION
    # -------------------------------------------------------------------------
    def apply_universal_headers(ws, title_text):
        ws.views.sheetView[0].showGridLines = False
        ws.freeze_panes = "A5"
        
        # Row 1 Title
        ws.merge_cells("A1:O1")
        ws["A1"] = title_text
        ws["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=COLOR_WHITE)
        ws["A1"].fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        
        # Setup layouts for Row 2,3,4
        ws.merge_cells("A2:E3")
        ws["A2"] = "Target = D+0 % to be above 90%, DSS Usage to be 100%"
        
        ws.merge_cells("F2:K2")
        ws["F2"] = f"Delivery Transit Analysis from {cumulative_range}"
        
        ws.merge_cells("L2:M3")
        ws["L2"] = f"Delivery Productivity % {daily_date}"
        
        ws.merge_cells("N2:O2")
        ws["N2"] = "DSS Usage"
        
        ws.merge_cells("F3:G3")
        ws["F3"] = "All Products"
        
        ws.merge_cells("H3:I3")
        ws["H3"] = "Documents"
        
        ws.merge_cells("J3:K3")
        ws["J3"] = "Parcel"
        
        ws.merge_cells("N3:N4")
        ws["N3"] = f"From {dss_range} (%)"
        
        ws.merge_cells("O3:O4")
        ws["O3"] = f"On {daily_date} (%)"
        
        col_labels = {
            1: "Sr No.", 2: "Sub Division Name", 3: "Sub Office Name", 4: "Office Name", 5: "Office Type",
            6: "Received", 7: "D+0 %", 8: "Received", 9: "D+0 %", 10: "Received", 11: "D+0 %",
            12: "Received", 13: "D+0 %", 14: "", 15: ""
        }
        for col_idx, label in col_labels.items():
            if label != "":
                ws.cell(row=4, column=col_idx, value=label)
                
        # Style Header Rows 2-4
        for r in [2, 3, 4]:
            ws.row_dimensions[r].height = 25
            for c in range(1, 16):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_NAVY)
                cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = header_cell_border

    def write_data_rows(ws, dataframe, is_sheet_1=False, only_bo=False, excl_bo=False):
        current_row = 5
        sub_divisions = sorted(dataframe['Sub Division'].unique())
        
        grand_totals = {
            'all_p_rec': 0, 'all_p_del': 0, 'all_p_red': 0, 'all_p_ret': 0,
            'doc_rec': 0, 'doc_del': 0, 'doc_red': 0, 'doc_ret': 0,
            'par_rec': 0, 'par_del': 0, 'par_red': 0, 'par_ret': 0,
            'dp_rec': 0, 'dp_del': 0, 'dp_red': 0, 'dp_ret': 0,
            'dss_cum_num': 0, 'dss_cum_den': 0,
            'dss_day_num': 0, 'dss_day_den': 0
        }
        
        for sdn in sub_divisions:
            df_sdn = dataframe[dataframe['Sub Division'] == sdn].sort_values(by=['Sub Office', 'Branch Office'])
            if df_sdn.empty:
                continue
                
            sr_no = 1
            is_first_row_of_block = True
            
            sub_totals = {
                'all_p_rec': 0, 'all_p_del': 0, 'all_p_red': 0, 'all_p_ret': 0,
                'doc_rec': 0, 'doc_del': 0, 'doc_red': 0, 'doc_ret': 0,
                'par_rec': 0, 'par_del': 0, 'par_red': 0, 'par_ret': 0,
                'dp_rec': 0, 'dp_del': 0, 'dp_red': 0, 'dp_ret': 0,
                'dss_cum_num': 0, 'dss_cum_den': 0,
                'dss_day_num': 0, 'dss_day_den': 0
            }
            
            for _, row in df_sdn.iterrows():
                ws.row_dimensions[current_row].height = 18
                bg_color = COLOR_WHITE if current_row % 2 == 0 else COLOR_LIGHT_GREY
                row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                
                # Metrics parsing
                all_p_rec = row.get('all_prod_Received', 0)
                all_p_del = row.get('all_prod_D0 Delivered', 0)
                all_p_red = row.get('all_prod_D0 Redirected', 0)
                all_p_ret = row.get('all_prod_D0 Returned', 0)
                
                reg_l_rec = row.get('reg_let_Received', 0)
                reg_l_del = row.get('reg_let_D0 Delivered', 0)
                reg_l_red = row.get('reg_let_D0 Redirected', 0)
                reg_l_ret = row.get('reg_let_D0 Returned', 0)
                
                spd_l_rec = row.get('spd_let_Received', 0)
                spd_l_del = row.get('spd_let_D0 Delivered', 0)
                spd_l_red = row.get('spd_let_D0 Redirected', 0)
                spd_l_ret = row.get('spd_let_D0 Returned', 0)
                
                reg_p_rec = row.get('reg_par_Received', 0)
                reg_p_del = row.get('reg_par_D0 Delivered', 0)
                reg_p_red = row.get('reg_par_D0 Redirected', 0)
                reg_p_ret = row.get('reg_par_D0 Returned', 0)
                
                spd_p_rec = row.get('spd_par_Received', 0)
                spd_p_del = row.get('spd_par_D0 Delivered', 0)
                spd_p_red = row.get('spd_par_D0 Redirected', 0)
                spd_p_ret = row.get('spd_par_D0 Returned', 0)
                
                dp_rec = row.get('dp_invoice-count', 0)
                dp_del = row.get('dp_delivery-count', 0)
                dp_red = row.get('dp_redirection-count', 0)
                dp_ret = row.get('dp_return-count', 0)
                
                dss_cum_num = row.get('dss_cum_total_dss_art_count', 0)
                dss_cum_den = row.get('dss_cum_total_pdm_art_count', 0)
                
                dss_day_num = row.get('dss_day_total_dss_art_count', 0)
                dss_day_den = row.get('dss_day_total_pdm_art_count', 0)
                
                # Composite math
                doc_rec = reg_l_rec + spd_l_rec
                doc_del = reg_l_del + spd_l_del
                doc_red = reg_l_red + spd_l_red
                doc_ret = reg_l_ret + spd_l_ret
                
                par_rec = reg_p_rec + spd_p_rec
                par_del = reg_p_del + spd_p_del
                par_red = reg_p_red + spd_p_red
                par_ret = reg_p_ret + spd_p_ret
                
                # Accumulate subtotals
                for key, val in zip(
                    ['all_p_rec', 'all_p_del', 'all_p_red', 'all_p_ret', 'doc_rec', 'doc_del', 'doc_red', 'doc_ret', 'par_rec', 'par_del', 'par_red', 'par_ret', 'dp_rec', 'dp_del', 'dp_red', 'dp_ret', 'dss_cum_num', 'dss_cum_den', 'dss_day_num', 'dss_day_den'],
                    [all_p_rec, all_p_del, all_p_red, all_p_ret, doc_rec, doc_del, doc_red, doc_ret, par_rec, par_del, par_red, par_ret, dp_rec, dp_del, dp_red, dp_ret, dss_cum_num, dss_cum_den, dss_day_num, dss_day_den]
                ):
                    sub_totals[key] += val
                    grand_totals[key] += val
                
                # Calculated final model metrics for row
                m_all_prod = safe_div_model_b(all_p_del, all_p_red, all_p_ret, all_p_rec)
                m_doc = safe_div_model_b(doc_del, doc_red, doc_ret, doc_rec) if (reg_l_rec > 0 or spd_l_rec > 0) else "-"
                m_par = safe_div_model_b(par_del, par_red, par_ret, par_rec) if (reg_p_rec > 0 or spd_p_rec > 0) else "-"
                m_dp = safe_div_model_b(dp_del, dp_red, dp_ret, dp_rec)
                m_dss_cum = safe_div_model_a(dss_cum_num, dss_cum_den)
                m_dss_day = safe_div_model_a(dss_day_num, dss_day_den)
                
                row_values = {
                    1: sr_no,
                    2: sdn if is_first_row_of_block else "",
                    3: row['Sub Office'],
                    4: row['Branch Office'],
                    5: row['office-type-code'],
                    6: format_cell_value(all_p_rec),
                    7: format_cell_value(m_all_prod),
                    8: format_cell_value(doc_rec),
                    9: format_cell_value(m_doc),
                    10: format_cell_value(par_rec),
                    11: format_cell_value(m_par),
                    12: format_cell_value(dp_rec),
                    13: format_cell_value(m_dp),
                    14: format_cell_value(m_dss_cum),
                    15: format_cell_value(m_dss_day)
                }
                
                for c_idx, val in row_values.items():
                    cell = ws.cell(row=current_row, column=c_idx, value=val)
                    cell.font = Font(name=FONT_NAME, size=9)
                    cell.fill = row_fill
                    cell.border = data_cell_border
                    
                    # Alignment and formatting
                    if c_idx in [1, 5]:
                        cell.alignment = Alignment(horizontal="center")
                    elif c_idx >= 6:
                        cell.alignment = Alignment(horizontal="right")
                        if isinstance(val, float):
                            cell.number_format = "0.0" if c_idx in [7, 9, 11, 13, 14, 15] else "#,##0"
                            
                        # Alert validations
                        if val != "-":
                            if c_idx in [7, 9, 11, 13] and val < 90.0:
                                cell.fill = PatternFill(start_color=COLOR_ALERT_FILL, end_color=COLOR_ALERT_FILL, fill_type="solid")
                                cell.font = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_ALERT_FONT)
                            elif c_idx in [14, 15] and val < 100.0:
                                cell.fill = PatternFill(start_color=COLOR_ALERT_FILL, end_color=COLOR_ALERT_FILL, fill_type="solid")
                                cell.font = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_ALERT_FONT)
                    else:
                        cell.alignment = Alignment(horizontal="left")
                
                sr_no += 1
                is_first_row_of_block = False
                current_row += 1
                
            # Write Subtotal Row (Ignore Sheet 1)
            if not is_sheet_1:
                ws.row_dimensions[current_row].height = 20
                sub_fill = PatternFill(start_color=COLOR_SUBTOTAL, end_color=COLOR_SUBTOTAL, fill_type="solid")
                sub_font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_NAVY)
                
                st_all_prod = safe_div_model_b(sub_totals['all_p_del'], sub_totals['all_p_red'], sub_totals['all_p_ret'], sub_totals['all_p_rec'])
                st_doc = safe_div_model_b(sub_totals['doc_del'], sub_totals['doc_red'], sub_totals['doc_ret'], sub_totals['doc_rec'])
                st_par = safe_div_model_b(sub_totals['par_del'], sub_totals['par_red'], sub_totals['par_ret'], sub_totals['par_rec'])
                st_dp = safe_div_model_b(sub_totals['dp_del'], sub_totals['dp_red'], sub_totals['dp_ret'], sub_totals['dp_rec'])
                st_dss_cum = safe_div_model_a(sub_totals['dss_cum_num'], sub_totals['dss_cum_den'])
                st_dss_day = safe_div_model_a(sub_totals['dss_day_num'], sub_totals['dss_day_den'])
                
                ws.cell(row=current_row, column=1, value="")
                ws.cell(row=current_row, column=2, value=sdn)
                
                if excl_bo:
                    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
                    ws.cell(row=current_row, column=3, value=f"{sdn} (Total)")
                elif only_bo:
                    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
                    ws.cell(row=current_row, column=3, value=f"{sdn} Total")
                
                st_values = {
                    5: "" if excl_bo else None,
                    6: format_cell_value(sub_totals['all_p_rec']), 7: format_cell_value(st_all_prod),
                    8: format_cell_value(sub_totals['doc_rec']), 9: format_cell_value(st_doc),
                    10: format_cell_value(sub_totals['par_rec']), 11: format_cell_value(st_par),
                    12: format_cell_value(sub_totals['dp_rec']), 13: format_cell_value(st_dp),
                    14: format_cell_value(st_dss_cum), 15: format_cell_value(st_dss_day)
                }
                
                for col_idx in range(1, 16):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.fill = sub_fill
                    cell.font = sub_font
                    cell.border = header_cell_border
                    
                    if col_idx in st_values and st_values[col_idx] is not None:
                        cell.value = st_values[col_idx]
                    
                    if col_idx in [1, 5]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx >= 6:
                        cell.alignment = Alignment(horizontal="right")
                        val = cell.value
                        if isinstance(val, float):
                            cell.number_format = "0.0" if col_idx in [7, 9, 11, 13, 14, 15] else "#,##0"
                    else:
                        if col_idx == 3:
                            cell.alignment = Alignment(horizontal="left")
                            
                current_row += 1
                
        # Write Grand Total Row (Ignore Sheet 1)
        if not is_sheet_1:
            ws.row_dimensions[current_row].height = 22
            gt_fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
            gt_font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_WHITE)
            
            gt_all_prod = safe_div_model_b(grand_totals['all_p_del'], grand_totals['all_p_red'], grand_totals['all_p_ret'], grand_totals['all_p_rec'])
            gt_doc = safe_div_model_b(grand_totals['doc_del'], grand_totals['doc_red'], grand_totals['doc_ret'], grand_totals['doc_rec'])
            gt_par = safe_div_model_b(grand_totals['par_del'], grand_totals['par_red'], grand_totals['par_ret'], grand_totals['par_rec'])
            gt_dp = safe_div_model_b(grand_totals['dp_del'], grand_totals['dp_red'], grand_totals['dp_ret'], grand_totals['dp_rec'])
            gt_dss_cum = safe_div_model_a(grand_totals['dss_cum_num'], grand_totals['dss_cum_den'])
            gt_dss_day = safe_div_model_a(grand_totals['dss_day_num'], grand_totals['dss_day_den'])
            
            if excl_bo:
                ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
                ws.cell(row=current_row, column=3, value="Karad Division (Total)")
            elif only_bo:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                ws.cell(row=current_row, column=1, value="Karad Division Total")
                
            gt_values = {
                6: format_cell_value(grand_totals['all_p_rec']), 7: format_cell_value(gt_all_prod),
                8: format_cell_value(grand_totals['doc_rec']), 9: format_cell_value(gt_doc),
                10: format_cell_value(grand_totals['par_rec']), 11: format_cell_value(gt_par),
                12: format_cell_value(grand_totals['dp_rec']), 13: format_cell_value(gt_dp),
                14: format_cell_value(gt_dss_cum), 15: format_cell_value(gt_dss_day)
            }
            
            for col_idx in range(1, 16):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill = gt_fill
                cell.font = gt_font
                cell.border = header_cell_border
                
                if col_idx in gt_values:
                    cell.value = gt_values[col_idx]
                
                if col_idx >= 6:
                    cell.alignment = Alignment(horizontal="right")
                    val = cell.value
                    if isinstance(val, float):
                        cell.number_format = "0.0" if col_idx in [7, 9, 11, 13, 14, 15] else "#,##0"
                else:
                    if (excl_bo and col_idx == 3) or (only_bo and col_idx == 1):
                        cell.alignment = Alignment(horizontal="left")

    # -------------------------------------------------------------------------
    # SHEET 1: Raw Data Sheet Office wise
    # -------------------------------------------------------------------------
    ws1 = wb.create_sheet(title="Raw Data Sheet Office wise")
    apply_universal_headers(ws1, f"MMU Report {daily_date} (All Offices)")
    write_data_rows(ws1, df_final, is_sheet_1=True)
    for col_letter, width in COL_WIDTHS_1235.items():
        ws1.column_dimensions[col_letter].width = width

    # -------------------------------------------------------------------------
    # SHEET 2: Excluding B.Os
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Excluding B.Os")
    apply_universal_headers(ws2, f"MMU Report {daily_date} (Excluding BOs)")
    df_ex_bo = df_final[df_final['office-type-code'].isin(['HPO', 'SPO'])]
    write_data_rows(ws2, df_ex_bo, is_sheet_1=False, excl_bo=True)
    for col_letter, width in COL_WIDTHS_1235.items():
        ws2.column_dimensions[col_letter].width = width

    # -------------------------------------------------------------------------
    # SHEET 3: Only B.Os
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet(title="Only B.Os")
    apply_universal_headers(ws3, f"MMU Report {daily_date} (Only BOs)")
    df_only_bo = df_final[df_final['office-type-code'] == 'BPO']
    write_data_rows(ws3, df_only_bo, is_sheet_1=False, only_bo=True)
    for col_letter, width in COL_WIDTHS_1235.items():
        ws3.column_dimensions[col_letter].width = width

    # -------------------------------------------------------------------------
    # SHEET 4: Sub Division wise Summary
    # -------------------------------------------------------------------------
    ws4 = wb.create_sheet(title="Sub Division wise Summary")
    ws4.views.sheetView[0].showGridLines = False
    ws4.freeze_panes = "A6"
    
    # Layout configuration
    ws4.merge_cells("A1:N1")
    ws4["A1"] = f"MMU Report {daily_date} (SDn wise Summary)"
    ws4["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=COLOR_WHITE)
    ws4["A1"].fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28
    
    ws4.merge_cells("A2:B4")
    ws4["A2"] = "Target D+0 % Above 90%; DSS Usage 100%"
    
    ws4.merge_cells("C2:G2")
    ws4["C2"] = f"Delivery Transit from {cumulative_range}"
    
    ws4.merge_cells("H2:H3")
    ws4["H2"] = f"DSS Usage from {dss_range}"
    
    ws4.merge_cells("I2:M2")
    ws4["I2"] = f"Delivery Transit from {cumulative_range}"
    
    ws4.merge_cells("N2:N3")
    ws4["N2"] = f"DSS Usage from {dss_range}"
    
    ws4.merge_cells("C3:G3")
    ws4["C3"] = "Excluding B.Os"
    
    ws4.merge_cells("I3:M3")
    ws4["I3"] = "Only B.Os"
    
    ws4.merge_cells("D4:E4")
    ws4["D4"] = "Documents"
    ws4.merge_cells("F4:G4")
    ws4["F4"] = "Parcel"
    ws4.merge_cells("J4:K4")
    ws4["J4"] = "Documents"
    ws4.merge_cells("L4:M4")
    ws4["L4"] = "Parcel"
    
    col_labels_4 = {
        1: "Sr. No.", 2: "Sub Division", 3: "Total Offices", 4: "Received", 5: "D+0%",
        6: "Received", 7: "D+0%", 8: "DSS %", 9: "Total Offices", 10: "Received",
        11: "D+0%", 12: "Received", 13: "D+0%", 14: "DSS %"
    }
    for c_idx, label in col_labels_4.items():
        ws4.cell(row=5, column=c_idx, value=label)
        
    for r in [2, 3, 4, 5]:
        ws4.row_dimensions[r].height = 22 if r == 5 else 25
        for c in range(1, 15):
            cell = ws4.cell(row=r, column=c)
            cell.font = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_NAVY)
            cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_cell_border

    # Build aggregates for Sheet 4
    sub_divisions_list = sorted(df_final['Sub Division'].unique())
    c_row = 6
    s4_sr = 1
    
    div_totals = {
        'ex_cnt': 0, 'ex_doc_rec': 0, 'ex_doc_del': 0, 'ex_doc_red': 0, 'ex_doc_ret': 0,
        'ex_par_rec': 0, 'ex_par_del': 0, 'ex_par_red': 0, 'ex_par_ret': 0,
        'ex_dss_num': 0, 'ex_dss_den': 0,
        'bo_cnt': 0, 'bo_doc_rec': 0, 'bo_doc_del': 0, 'bo_doc_red': 0, 'bo_doc_ret': 0,
        'bo_par_rec': 0, 'bo_par_del': 0, 'bo_par_red': 0, 'bo_par_ret': 0,
        'bo_dss_num': 0, 'bo_dss_den': 0
    }
    
    for sdn in sub_divisions_list:
        ws4.row_dimensions[c_row].height = 18
        bg_color = COLOR_WHITE if c_row % 2 == 0 else COLOR_LIGHT_GREY
        s4_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Subsets
        df_sdn_ex = df_ex_bo[df_ex_bo['Sub Division'] == sdn]
        df_sdn_bo = df_only_bo[df_only_bo['Sub Division'] == sdn]
        
        # Counts from static config map (pre-calculated during initial filters)
        ex_cnt = counts_map[sdn]['ex']
        bo_cnt = counts_map[sdn]['bo']
        
        # Aggregate logic
        def sum_comp(df):
            doc_rec = (df['reg_let_Received'].sum() + df['spd_let_Received'].sum())
            doc_del = (df['reg_let_D0 Delivered'].sum() + df['spd_let_D0 Delivered'].sum())
            doc_red = (df['reg_let_D0 Redirected'].sum() + df['spd_let_D0 Redirected'].sum())
            doc_ret = (df['reg_let_D0 Returned'].sum() + df['spd_let_D0 Returned'].sum())
            
            par_rec = (df['reg_par_Received'].sum() + df['spd_par_Received'].sum())
            par_del = (df['reg_par_D0 Delivered'].sum() + df['spd_par_D0 Delivered'].sum())
            par_red = (df['reg_par_D0 Redirected'].sum() + df['spd_par_D0 Redirected'].sum())
            par_ret = (df['reg_par_D0 Returned'].sum() + df['spd_par_D0 Returned'].sum())
            
            dss_num = df['dss_cum_total_dss_art_count'].sum()
            dss_den = df['dss_cum_total_pdm_art_count'].sum()
            
            return doc_rec, doc_del, doc_red, doc_ret, par_rec, par_del, par_red, par_ret, dss_num, dss_den

        ex_m = sum_comp(df_sdn_ex)
        bo_m = sum_comp(df_sdn_bo)
        
        # Accumulate Division
        div_totals['ex_cnt'] += ex_cnt
        div_totals['bo_cnt'] += bo_cnt
        for idx, k in enumerate(['doc_rec', 'doc_del', 'doc_red', 'doc_ret', 'par_rec', 'par_del', 'par_red', 'par_ret', 'dss_num', 'dss_den']):
            div_totals['ex_' + k] += ex_m[idx]
            div_totals['bo_' + k] += bo_m[idx]
            
        # Percents
        p_ex_doc = safe_div_model_b(ex_m[1], ex_m[2], ex_m[3], ex_m[0])
        p_ex_par = safe_div_model_b(ex_m[5], ex_m[6], ex_m[7], ex_m[4])
        p_ex_dss = safe_div_model_a(ex_m[8], ex_m[9])
        
        p_bo_doc = safe_div_model_b(bo_m[1], bo_m[2], bo_m[3], bo_m[0])
        p_bo_par = safe_div_model_b(bo_m[5], bo_m[6], bo_m[7], bo_m[4])
        p_bo_dss = safe_div_model_a(bo_m[8], bo_m[9])
        
        s4_vals = {
            1: s4_sr, 2: sdn, 3: format_cell_value(ex_cnt),
            4: format_cell_value(ex_m[0]), 5: format_cell_value(p_ex_doc),
            6: format_cell_value(ex_m[4]), 7: format_cell_value(p_ex_par),
            8: format_cell_value(p_ex_dss), 9: format_cell_value(bo_cnt),
            10: format_cell_value(bo_m[0]), 11: format_cell_value(p_bo_doc),
            12: format_cell_value(bo_m[4]), 13: format_cell_value(p_bo_par),
            14: format_cell_value(p_bo_dss)
        }
        
        for col_idx, val in s4_vals.items():
            cell = ws4.cell(row=c_row, column=col_idx, value=val)
            cell.font = Font(name=FONT_NAME, size=9)
            cell.fill = s4_fill
            cell.border = data_cell_border
            
            if col_idx in [1, 3, 9]:
                cell.alignment = Alignment(horizontal="center")
                if col_idx in [3, 9] and isinstance(val, float):
                    cell.number_format = "#,##0"
            elif col_idx >= 4:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, float):
                    cell.number_format = "0.0" if col_idx in [5, 7, 8, 11, 13, 14] else "#,##0"
                if val != "-":
                    if col_idx in [5, 7, 11, 13] and val < 90.0:
                        cell.fill = PatternFill(start_color=COLOR_ALERT_FILL, end_color=COLOR_ALERT_FILL, fill_type="solid")
                        cell.font = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_ALERT_FONT)
                    elif col_idx in [8, 14] and val < 100.0:
                        cell.fill = PatternFill(start_color=COLOR_ALERT_FILL, end_color=COLOR_ALERT_FILL, fill_type="solid")
                        cell.font = Font(name=FONT_NAME, size=9, bold=True, color=COLOR_ALERT_FONT)
            else:
                cell.alignment = Alignment(horizontal="left")
                
        s4_sr += 1
        c_row += 1
        
    # Sheet 4 Grand Total
    ws4.row_dimensions[c_row].height = 22
    s4_gt_fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    s4_gt_font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_WHITE)
    
    gt_ex_doc = safe_div_model_b(div_totals['ex_doc_del'], div_totals['ex_doc_red'], div_totals['ex_doc_ret'], div_totals['ex_doc_rec'])
    gt_ex_par = safe_div_model_b(div_totals['ex_par_del'], div_totals['ex_par_red'], div_totals['ex_par_ret'], div_totals['ex_par_rec'])
    gt_ex_dss = safe_div_model_a(div_totals['ex_dss_num'], div_totals['ex_dss_den'])
    
    gt_bo_doc = safe_div_model_b(div_totals['bo_doc_del'], div_totals['bo_doc_red'], div_totals['bo_doc_ret'], div_totals['bo_doc_rec'])
    gt_bo_par = safe_div_model_b(div_totals['bo_par_del'], div_totals['bo_par_red'], div_totals['bo_par_ret'], div_totals['bo_par_rec'])
    gt_bo_dss = safe_div_model_a(div_totals['bo_dss_num'], div_totals['bo_dss_den'])
    
    ws4.merge_cells(start_row=c_row, start_column=1, end_row=c_row, end_column=2)
    ws4.cell(row=c_row, column=1, value="Karad Division Total")
    
    s4_gt_vals = {
        3: format_cell_value(div_totals['ex_cnt']), 4: format_cell_value(div_totals['ex_doc_rec']),
        5: format_cell_value(gt_ex_doc), 6: format_cell_value(div_totals['ex_par_rec']),
        7: format_cell_value(gt_ex_par), 8: format_cell_value(gt_ex_dss),
        9: format_cell_value(div_totals['bo_cnt']), 10: format_cell_value(div_totals['bo_doc_rec']),
        11: format_cell_value(gt_bo_doc), 12: format_cell_value(div_totals['bo_par_rec']),
        13: format_cell_value(gt_bo_par), 14: format_cell_value(gt_bo_dss)
    }
    
    for col_idx in range(1, 15):
        cell = ws4.cell(row=c_row, column=col_idx)
        cell.fill = s4_gt_fill
        cell.font = s4_gt_font
        cell.border = header_cell_border
        
        if col_idx in s4_gt_vals:
            cell.value = s4_gt_vals[col_idx]
            
        if col_idx >= 3:
            cell.alignment = Alignment(horizontal="right" if col_idx not in [3, 9] else "center")
            val = cell.value
            if isinstance(val, float):
                cell.number_format = "0.0" if col_idx in [5, 7, 8, 11, 13, 14] else "#,##0"
        else:
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
                
    for col_letter, width in COL_WIDTHS_4.items():
        ws4.column_dimensions[col_letter].width = width

    # -------------------------------------------------------------------------
    # SHEET 5: Defaulter Offices list
    # -------------------------------------------------------------------------
    ws5 = wb.create_sheet(title="Defaulter Offices list")
    apply_universal_headers(ws5, f"MMU Report {daily_date} (KPI missed Offices list)")
    
    # Identify defaulters directly dynamically
    defaulter_rows = []
    for _, row in df_final.iterrows():
        all_p_rec = row.get('all_prod_Received', 0)
        all_p_del = row.get('all_prod_D0 Delivered', 0)
        all_p_red = row.get('all_prod_D0 Redirected', 0)
        all_p_ret = row.get('all_prod_D0 Returned', 0)
        
        dss_cum_num = row.get('dss_cum_total_dss_art_count', 0)
        dss_cum_den = row.get('dss_cum_total_pdm_art_count', 0)
        
        m_all_prod = safe_div_model_b(all_p_del, all_p_red, all_p_ret, all_p_rec)
        m_dss_cum = safe_div_model_a(dss_cum_num, dss_cum_den)
        
        is_def = False
        if m_all_prod != "-" and m_all_prod < 90.0:
            is_def = True
        if m_dss_cum != "-" and m_dss_cum < 90.0:
            is_def = True
            
        if is_def:
            defaulter_rows.append(row)
            
    if defaulter_rows:
        df_def = pd.DataFrame(defaulter_rows).sort_values(by=['Sub Division', 'Branch Office'])
        write_data_rows(ws5, df_def, is_sheet_1=True)
    else:
        # If no defaulters write an empty placeholder styled row
        ws5.cell(row=5, column=1, value="No KPI Defaulter Offices Found.").font = Font(name=FONT_NAME, size=10, italic=True)
        
    for col_letter, width in COL_WIDTHS_1235.items():
        ws5.column_dimensions[col_letter].width = width

    # Write out in memory
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf

# -----------------------------------------------------------------------------
# STREAMLIT INTERFACE ENTRYPOINT
# -----------------------------------------------------------------------------
def main():
    st.set_page_config(page_title="MMU Report Generator", layout="centered")
    
    st.title("📊 MMU Report Generator — Karad Division")
    st.markdown("##### Department of Posts | Office of the Superintendent, Karad Division")
    st.markdown("---")
    
    # Load Master Directory Setup
    try:
        df_master = pd.read_csv(MASTER_FILE_PATH)
        # Verify exact column names match specifications
        req_cols = ["Division Office Name", "Sub Division", "Office ID", "Sub Office", "Branch Office", "office-type-code", "PIN"]
        for rc in req_cols:
            if rc not in df_master.columns:
                st.error(f"Critical Error: Master directory file is missing standard column architecture structure: '{rc}'")
                st.stop()
                
        # Clean master key data types right away
        df_master['Office ID'] = df_master['Office ID'].astype(str).str.strip()
        
        # Calculate dynamic display count metrics while excluding requested offices safely
        df_master_clean = df_master[
            ~df_master['Branch Office'].str.contains('shenawadi|yeralwadi', case=False, na=False)
        ]
        
        hpo_c = len(df_master_clean[df_master_clean['office-type-code'] == 'HPO'])
        spo_c = len(df_master_clean[df_master_clean['office-type-code'] == 'SPO'])
        bpo_c = len(df_master_clean[df_master_clean['office-type-code'] == 'BPO'])
        total_c = len(df_master_clean)
        
        st.success(f"Master Directory loaded: {total_c} active offices ({hpo_c} HPO + {spo_c} SPO + {bpo_c} BPO)")
        
        # Calculate static historical breakdown allocations for dynamic Sheet 4 office metrics counts
        counts_map = {}
        for sdn in df_master_clean['Sub Division'].unique():
            df_sub = df_master_clean[df_master_clean['Sub Division'] == sdn]
            counts_map[sdn] = {
                'ex': len(df_sub[df_sub['office-type-code'].isin(['HPO', 'SPO'])]),
                'bo': len(df_sub[df_sub['office-type-code'] == 'BPO'])
            }
            
    except FileNotFoundError:
        st.error(f"🚨 Master Directory file not found natively at: `{MASTER_FILE_PATH}`. Deployment execution halted.")
        st.stop()
    except Exception as e:
        st.error(f"Error initializing data pipeline configuration at startup: {e}")
        st.stop()

    # -------------------------------------------------------------------------
    # SECTION A: Cumulative Transit Files Layout UI
    # -------------------------------------------------------------------------
    st.write("### SECTION A — Cumulative Transit Files")
    col1, col2 = st.columns(2)
    
    with col1:
        f_all_prod = st.file_uploader("All Products CSV ❗", type=["csv"], key="all_prod")
        st.caption("✅ File Uploaded" if f_all_prod else "⬜ Pending Upload")
        
        f_reg_let = st.file_uploader("Registered Letter CSV", type=["csv"], key="reg_let")
        st.caption("✅ File Uploaded" if f_reg_let else "⬜ Pending Upload")
        
        f_spd_let = st.file_uploader("Speed Letter CSV", type=["csv"], key="spd_let")
        st.caption("✅ File Uploaded" if f_spd_let else "⬜ Pending Upload")
        
    with col2:
        f_reg_par = st.file_uploader("Registered Parcel CSV", type=["csv"], key="reg_par")
        st.caption("✅ File Uploaded" if f_reg_par else "⬜ Pending Upload")
        
        f_spd_par = st.file_uploader("Speed Parcel CSV", type=["csv"], key="spd_par")
        st.caption("✅ File Uploaded" if f_spd_par else "⬜ Pending Upload")

    # -------------------------------------------------------------------------
    # SECTION B: Daily Snapshots Layout UI
    # -------------------------------------------------------------------------
    st.write("### SECTION B — Daily Snapshot")
    col3, col4 = st.columns(2)
    
    with col3:
        f_dp = st.file_uploader("Delivery Productivity CSV ❗", type=["csv"], key="dp")
        st.caption("✅ File Uploaded" if f_dp else "⬜ Pending Upload")
        
    with col4:
        f_dss_day = st.file_uploader("Daily DSS Usage CSV", type=["csv"], key="dss_day")
        st.caption("✅ File Uploaded" if f_dss_day else "⬜ Pending Upload")

    # -------------------------------------------------------------------------
    # SECTION C: Cumulative DSS Usage UI
    # -------------------------------------------------------------------------
    st.write("### SECTION C — Cumulative DSS Usage")
    f_dss_cum = st.file_uploader("DSS Usage Cumulative Range CSV (NOT a daily file)", type=["csv"], key="dss_cum")
    st.caption("✅ File Uploaded" if f_dss_cum else "⬜ Pending Upload")

    # -------------------------------------------------------------------------
    # SECTION D: Date Configuration Layout UI
    # -------------------------------------------------------------------------
    st.write("### SECTION D — Date Configuration")
    col5, col6, col7 = st.columns(3)
    
    with col5:
        today_str = datetime.date.today().strftime("%d.%m.%Y")
        inp_daily_date = st.text_input("Report Date (DD.MM.YYYY)", value=today_str)
    with col6:
        inp_cum_range = st.text_input("Transit Cumulative Range", value="01.05.2026 to 22.05.2026")
    with col7:
        inp_dss_range = st.text_input("DSS Cumulative Range", value="01.05.2026 to 30.05.2026")

    st.write("---")

    # -------------------------------------------------------------------------
    # SECTION E: Generate Report Handling Orchestrator Execution Pipeline
    # -------------------------------------------------------------------------
    if st.button("⚙️ Generate MMU Report", use_container_width=True):
        missing_reqs = []
        if not f_all_prod:
            missing_reqs.append("All Products CSV")
        if not f_dp:
            missing_reqs.append("Delivery Productivity CSV")
            
        if missing_reqs:
            st.error(f"Cannot generate report. Missing required files: {', '.join(missing_reqs)}")
            return
            
        # Logging engine deployment
        with st.status("📋 Processing Log", expanded=True) as status:
            try:
                st.write("Reading inputs data...")
                
                # Base processing mapping pipeline function helper
                def parse_inmemory_csv(uploaded, prefix, drop_summary=False):
                    if uploaded is None:
                        return None
                    df = pd.read_csv(io.BytesIO(uploaded.read()))
                    
                    # Discover Key column safely
                    oid_col = None
                    for c in df.columns:
                        if str(c).strip().lower() in ['office_id', 'office-id']:
                            oid_col = c
                            break
                    if oid_col is None:
                        raise ValueError(f"Could not resolve key structural index element join matching key identity attribute map parameters block inside input data frame configurations template.")
                    
                    if drop_summary:
                        df = df[df[oid_col].astype(str).str.strip() != '0']
                        
                    df['join_key'] = df[oid_col].astype(str).str.strip()
                    
                    # Rename columns to avoid metric naming overwrites
                    rename_map = {}
                    for c in df.columns:
                        if c != 'join_key' and c != oid_col:
                            rename_map[c] = f"{prefix}_{c}"
                    df = df.rename(columns=rename_map)
                    
                    # Keep cleanly targeted calculation parameters attributes
                    keep_cols = ['join_key'] + [v for k, v in rename_map.items()]
                    return df[keep_cols]

                # Parse files data frames schemas objects mappings
                df_ap_parsed = parse_inmemory_csv(f_all_prod, "all_prod")
                df_rl_parsed = parse_inmemory_csv(f_reg_let, "reg_let")
                df_sl_parsed = parse_inmemory_csv(f_spd_let, "spd_let")
                df_rp_parsed = parse_inmemory_csv(f_reg_par, "reg_par")
                df_sp_parsed = parse_inmemory_csv(f_spd_par, "spd_par")
                df_dp_parsed = parse_inmemory_csv(f_dp, "dp")
                df_dc_parsed = parse_inmemory_csv(f_dss_cum, "dss_cum", drop_summary=True)
                df_dd_parsed = parse_inmemory_csv(f_dss_day, "dss_day", drop_summary=True)
                
                st.write("Applying master filters boundaries mappings...")
                # Start combining directly with base master records entries mapping
                df_working = df_master_clean.copy()
                df_working['join_key'] = df_working['Office ID']
                
                # Successive sequential lookup joints pipelines safely maps
                all_inputs = [
                    df_ap_parsed, df_rl_parsed, df_sl_parsed, df_rp_parsed,
                    df_sp_parsed, df_dp_parsed, df_dc_parsed, df_dd_parsed
                ]
                
                for df_in in all_inputs:
                    if df_in is not None:
                        df_working = pd.merge(df_working, df_in, on='join_key', how='left')
                        
                # Fill na targets fields allocations records safely
                df_working = df_working.fillna(0)
                
                st.write("Executing styling and openpyxl layout generation engines modules...")
                excel_buffer = generate_excel_report(
                    df_working, counts_map, inp_daily_date, inp_cum_range, inp_dss_range
                )
                
                status.update(label="Report Generation Completed Successfully!", state="complete", expanded=False)
                
                # Deliver finalized attachment object model parameters to Streamlit UI window download element block layout frame
                out_filename = f"MMU Report {inp_daily_date} (All Offices).xlsx"
                st.download_button(
                    label="📥 Download MMU Excel Report File",
                    data=excel_buffer,
                    file_name=out_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            except Exception as ex:
                status.update(label="Critical Exception Error Intercepted During Generation!", state="error")
                st.error("Traceback error validation dump output:")
                st.code(traceback.format_exc())

if __name__ == "__main__":
    main()
